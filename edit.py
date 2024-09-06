import sqlite3
import pandas as pd
from openpyxl import load_workbook

file_path = r'D:\Downloads\ps.xlsx'
sheet_name = 'template'

df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
excel = load_workbook(file_path, rich_text=True)
excel_sheet = excel[sheet_name]

conn = sqlite3.connect('order_details.db')
cursor = conn.cursor()


def identical(common_sku: str):
    cursor.execute('SELECT identical_sku FROM identical WHERE sku = ?', (common_sku,))
    identical_sku = cursor.fetchone()
    return identical_sku[0].strip() if identical_sku else common_sku.strip()


for row in excel_sheet.iter_rows(min_row=2, max_row=df.shape[0] + 2, min_col=1, max_col=excel_sheet.max_column):
    print(row[0].row, row[8].value)
    seller_sku = identical(row[8].value.split('-')[1].split('_')[0])
    price = float(row[11].value)
    if seller_sku.lower() in ['pink', 'black', 'aqua']:
        row[6].value = 'active'  # status
        row[9].value = 300  # stock
        if price < 330:
            row[11].value = 350  # price
excel.save(r'D:\Downloads\price stock result.xlsx')
