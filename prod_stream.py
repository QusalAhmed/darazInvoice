import glob
import io
import re
import math
import os
import sqlite3
import time
import requests
import winsound
import imagehash
import xml.etree.ElementTree as Et
from PIL import Image, ImageDraw, ImageFont
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from selenium import webdriver
from selenium.common import JavascriptException
from selenium.common.exceptions import (NoSuchElementException, TimeoutException, StaleElementReferenceException)
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager

# Local import
from colors import *
from daraz_api import DarazAPI
from gimini import model

# Variable
triumph = 0
directory = r'D:\Downloads\New Product'
image_hash_dict = {}

# Database setup
conn = sqlite3.connect('order_details.db')
cursor = conn.cursor()


def is_image(file_name):
    return file_name.endswith('.jpg') or file_name.endswith('.png') or file_name.endswith('.jfif') or \
        file_name.endswith('.jpeg') or file_name.endswith('.webp')


def save_sheet(sheet):
    while True:
        try:
            sheet.save(os.path.join(directory, 'product_info.xlsx'))
            break
        except PermissionError:
            print(RED + 'Close the file' + RESET + ' product_info.xlsx')
            winsound.PlaySound('SystemQuestion', winsound.SND_ALIAS)
            time.sleep(1)


def add_watermark(image_src, watermark_text):
    # Open the original image
    original = Image.open(image_src).convert("RGBA")
    # Make the image editable
    txt = Image.new("RGBA", original.size)
    # Load a font
    font = ImageFont.truetype("HARNGTON.TTF", original.height * 0.05)
    # Initialize ImageDraw
    draw = ImageDraw.Draw(txt)
    # Add text to image
    text_size = max(len(watermark_text), 12)
    draw.text((original.width - original.height * 0.025 * text_size - 10, original.height * 0.9),
              watermark_text, fill=(255, 255, 255, 255), font=font, stroke_width=3, stroke_fill=(0, 0, 0, 255), )
    # Combine original image with watermark
    watermarked = Image.alpha_composite(original, txt)
    # watermarked.convert("RGB").save((os.path.join(directory, image_src.split('.')[0] + '_watermarked.png')), "PNG")
    packet = io.BytesIO()
    watermarked.convert("RGB").save(packet, format='PNG')
    image_size = packet.getbuffer().nbytes
    if image_size > 3145728:
        print(RED + 'Watermark image size exceeded' + RESET)
        packet = io.BytesIO()
        watermarked.convert("RGB").save(packet, format='JPEG')
    # Return the final image
    return packet.getvalue()


def is_duplicate(image_hash, cutoff=20):
    global image_hash_dict

    for hash_value in image_hash_dict:
        if image_hash - hash_value <= cutoff:
            print(image_hash - hash_value, end=' ')
            return hash_value
    return False


def download_image(img_link):
    global image_hash_dict
    img_link = img_link.split('_')[0]
    # Check if image already downloaded
    if os.path.exists(os.path.join(directory, img_link.split('/')[-1])):
        print('Image already downloaded')
        return
    download_response = requests.get(img_link)
    if download_response.status_code == 200:
        # Get image dimension
        img = Image.open(io.BytesIO(download_response.content))
        img_width, img_height = img.size
        if img_width > 330 and img_height > 330:
            new_file_name = img_link.split('/')[-1]
            with open(os.path.join(directory, new_file_name), 'wb') as file:
                file.write(download_response.content)
                # Check successful writen
                while True:
                    if os.path.exists(os.path.join(directory, new_file_name)):
                        break
                    time.sleep(1)
            # Check for duplicate image
            new_image = Image.open(os.path.join(directory, new_file_name))
            gray_image = new_image.convert('L')
            gray_image = gray_image.resize((1000, 1000))
            hash_value = imagehash.whash(gray_image, hash_size=16)
            old_image_hash = is_duplicate(hash_value)
            if old_image_hash:
                print(GREEN + 'Duplicate image found' + RESET, end=' ')
                new_file_size = os.path.getsize(os.path.join(directory, new_file_name))
                old_file_name = image_hash_dict[old_image_hash]
                print(new_file_name, old_file_name)
                old_file_size = os.path.getsize(os.path.join(directory, old_file_name))
                if new_file_size > old_file_size:
                    os.remove(os.path.join(directory, old_file_name))
                    image_hash_dict.pop(old_image_hash)
                    image_hash_dict.update({hash_value: new_file_name})
                else:
                    os.remove(os.path.join(directory, new_file_name))
            else:
                image_hash_dict.update({hash_value: new_file_name})


def full_source(file_name):
    source = glob.glob(os.path.join(directory, f'{file_name}.*'))
    if len(source) > 1:
        print(RED + 'Multiple files found' + RESET)
    return source[0]


def get_image_url(image_src, online_image_dict):
    if '.' not in image_src:
        image_src = full_source(image_src)
    base_image_url = os.path.basename(image_src)
    return online_image_dict[os.path.splitext(base_image_url)[0]]


def square_image(image_src, daraz_api, image_dict):
    image_src = full_source(image_src)
    packet = io.BytesIO()
    # Open the image file
    img = Image.open(image_src)

    # Calculate the size difference
    width, height = img.size
    if width == height:
        return get_image_url(image_src, image_dict)

    # Create a square canvas
    new_size = min(width, height)
    new_img = Image.new('RGB', (new_size, new_size), (255, 255, 255))

    # Paste the original image on the square canvas, centered
    new_img.paste(img, ((new_size - width) // 2, (new_size - height) // 2))

    # Save the result
    new_img.save(packet, format='PNG')

    # Work with image size
    image_size = packet.getbuffer().nbytes
    if image_size > 3145728:
        print(BLUE + 'Watermark image size exceeded' + RESET)
        packet = io.BytesIO()
        new_img.convert("RGB").save(packet, format='JPEG')

    # Upload to daraz
    return daraz_api.upload_image(packet.getvalue())['data']['image']['url']


def extract_youtube_id(url):
    # Regular expression for matching YouTube URLs
    regex = r'(?:https?://)?(?:www\.)?(?:youtube\.com/(?:[^/]+/.*|(?:v|e(?:mbed)?|.+/d+|.*[?&]v=)|(?:.*[?&]v=))|youtu\.be/)([a-zA-Z0-9_-]{11})'

    match = re.search(regex, url)
    if match:
        return match.group(1)  # Return the video ID
    return input('Enter video ID: ')


def upload_product(root, sku_image_list, shop_name):
    print(YELLOW + f"Uploading product to {shop_name}" + RESET)
    cursor.execute('SELECT access_token FROM open_daraz WHERE shop_name = ?', (shop_name,))
    access_token = cursor.fetchone()[0]
    daraz_api = DarazAPI(access_token, shop_name)
    worksheet = load_workbook(os.path.join(directory, 'product_info.xlsx'))
    # Make images shop unique
    online_image_dict = {}
    for filename in os.listdir(directory):
        if filename.endswith(('png', 'jpg', 'jpeg', 'bmp', 'jfif', 'webp')):
            # Verify image dimensions
            image = Image.open(os.path.join(directory, filename))
            if image.width < 330 or image.height < 330:
                print(RED, 'Image dimension too small', RESET, filename)
                continue
            # Add watermark
            watermark_image = add_watermark(os.path.join(directory, filename), shop_name)
            response = daraz_api.upload_image(watermark_image)
            try:
                online_image_dict.update({os.path.splitext(filename)[0]: response['data']['image']['url']})
            except KeyError:
                print(RED, 'Image upload failed', RESET, filename, response)
                continue
            print(filename, response['data']['image']['url'])
    # Shop unique attribute
    # Main image
    root.find('Product').find('Images').clear()
    main_images = root.find('Product').find('Images')
    for serial in range(1, 9):
        # Check if image exists
        image_path = glob.glob(os.path.join(directory, f'{serial}.*'))
        if image_path:
            Et.SubElement(main_images, "Image").text = get_image_url(image_path[0], online_image_dict)
        else:
            continue

    # Name
    product_name_eng = worksheet['Product Name']['A1'].value
    product_name_bn = worksheet['Product Name']['A2'].value
    if product_name_bn is None:
        # Translate product_name_eng to Bengali
        product_name_bn = model.generate_content(
            [f":Translate {product_name_eng} to Bengali",
             "Only give translated text, don't add extra information"]).text

    qr = time.strftime("%Y%m%d%H%M%S").split('.')[0]
    root.find('Product').find('Attributes').find('name_en').text = product_name_eng + f' QR:{qr}'
    root.find('Product').find('Attributes').find('name').text = product_name_bn + f' QR:{qr}'

    # Description
    highlighted_description = worksheet['Description']['A1'].value
    normal_description = worksheet['Description']['B1'].value

    # Create description in html
    soup = BeautifulSoup('', 'html.parser')

    # Highlighted
    if highlighted_description:
        highlighted_paragraph = soup.new_tag('p')
        highlighted_paragraph['style'] = ("line-height:1.7;text-align:center;text-indent:0;margin-left:0;margin-top:0; \
                                                margin-bottom:0")
        soup.append(highlighted_paragraph)
        text_span = soup.new_tag('span')
        text_span['style'] = ("font-weight:bold;color:rgb(64, 64, 64);"
                              "background-color:rgb(255, 217, 102);font-size:12pt")
        text_span.string = highlighted_description
        highlighted_paragraph.append(text_span)

    paragraph = soup.new_tag('p')
    paragraph['style'] = "line-height:1.7;text-align:left;text-indent:0;margin-left:0;margin-top:0;margin-bottom:0"
    soup.append(paragraph)

    if normal_description is not None:
        text_span = soup.new_tag('span')
        text_span.append(BeautifulSoup(normal_description, 'html.parser'))
        paragraph.append(text_span)

    # Iterate through the images in the dictionary
    for image_source in online_image_dict:
        description_image = soup.new_tag('img')
        description_image['src'] = online_image_dict[image_source]
        description_image['style'] = "width:100%;display:inline;vertical-align:middle"
        paragraph.append(description_image)
    root.find('Product').find('Attributes').find('description').text = soup.decode_contents()
    root.find('Product').find('Attributes').find('description_en').text = soup.decode_contents()

    # White background image
    image_path = worksheet['Video URL']['B1'].value
    if image_path:
        root.find('Product').find('Attributes').find('promotion_whitebkg_image').text \
            = square_image(image_path, daraz_api, online_image_dict)
    else:
        root.find('Product').find('Attributes').find('promotion_whitebkg_image').text \
            = square_image('1', daraz_api, online_image_dict)
    # SKU image
    for index, sku_element in enumerate(root.find('Product').find('Skus').findall('Sku')):
        sku_element.find('Images').find('Image').text = get_image_url(sku_image_list[int(index)], online_image_dict)

    tree = Et.ElementTree(root)
    tree.write("library.xml", encoding="utf-8", xml_declaration=True)

    response = daraz_api.create_product(Et.tostring(root, encoding='utf-8').decode('utf-8'))
    print(YELLOW, shop_name, response, RESET)
    try:
        if response['code'] == 'ApiCallLimit':
            time.sleep(1)
            response = daraz_api.create_product(Et.tostring(root, encoding='utf-8').decode('utf-8'))
            print(YELLOW, shop_name, response, RESET)
    except KeyError:
        pass
    winsound.Beep(1000, 1000)
    # input('Press enter to continue')


def setup_directory():
    winsound.Beep(1000, 1000)
    if 'y' == input('Delete all files in the directory? (y/n): ').lower():
        for filename in os.listdir(directory):
            file_path = os.path.join(directory, filename)
            if file_path.endswith('.xlsx'):
                continue
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    os.rmdir(file_path)
                print('Successfully deleted %s' % file_path)
            except Exception as e:
                print(RED, 'Failed to delete %s. Reason: %s' % (file_path, e), RESET)


def setup_excel():
    # Create excel product info file if already not created
    if not os.path.exists(os.path.join(directory, 'product_info.xlsx')):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Video URL'
        ws.column_dimensions['A'].width = 100
        ws.column_dimensions['B'].width = 100

        wb.create_sheet('Highlights')
        wb['Highlights'].column_dimensions['A'].width = 100
        wb['Highlights'].alignment = Alignment(wrap_text=True)

        wb.create_sheet('Description')
        # Set column width
        for col in ['A', 'B']:
            wb['Description'].column_dimensions[col].width = 100
            wb['Description'].alignment = Alignment(wrap_text=True)

        wb.create_sheet('Category').column_dimensions['A'].width = 20
        wb['Category'].alignment = Alignment(horizontal='center', vertical='center')

        wb.create_sheet('Product Name').append([' '])
        wb['Product Name'].column_dimensions['A'].width = 100
        wb['Product Name'].column_dimensions['B'].width = 100

        wb.create_sheet('SKU Info').append(
            ['Color Family', 'Seller SKU', 'Price', 'Image URL', 'Dimensions', 'Weight', ])
        # Set column width
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            wb['SKU Info'].column_dimensions[col].width = 30
            wb['SKU Info'][col + '1'].alignment = Alignment(horizontal='center', vertical='center',
                                                            wrap_text=True)
            wb['SKU Info'][col + '1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        wb['SKU Info'].column_dimensions['D'].width = 50
        wb['SKU Info'].column_dimensions['C'].width = 20
        wb['SKU Info'].column_dimensions['B'].width = 40

        wb.create_sheet('Attributes').append(['Attribute', 'Value'])
        for col in ['A', 'B']:
            wb['Attributes'].column_dimensions[col].width = 50
            wb['Attributes'][col + '1'].alignment = Alignment(horizontal='center', vertical='center',
                                                              wrap_text=True)
            wb['Attributes'][col + '1'].fill = PatternFill(start_color="00FFFF", end_color="00FFFF",
                                                           fill_type="solid")
            # Set font size
            wb['Attributes'][col + '1'].font = Font(bold=True, size=14)

        # Save the file
        wb.save(os.path.join(directory, 'product_info.xlsx'))


def setup_driver():
    service = ChromeService(ChromeDriverManager().install())
    options = webdriver.ChromeOptions()
    # options.add_argument('--headless')
    options.add_argument('--disable-gpu')
    options.add_argument('--start-maximized')
    options.add_argument("--mute-audio")
    driver = webdriver.Chrome(options=options, service=service)
    driver.implicitly_wait(10)
    wait = WebDriverWait(driver, 20)
    mouse = webdriver.ActionChains(driver)
    return driver, wait, mouse


def scrap_handler():
    driver, wait, mouse = setup_driver()
    # Scrapping images
    driver.get('https://www.daraz.com.bd/')
    while True:
        try:
            if len(driver.window_handles) >= 2:
                driver.switch_to.window(driver.window_handles[1])
                WebDriverWait(driver, 30).until(
                    lambda l_driver: l_driver.execute_script('return document.readyState') == 'complete'
                )
            else:
                winsound.Beep(1000, 1000)
                product_url = input('Enter product URL or enter to exit: ').split('?')[0]
                if product_url == '':
                    if len(driver.window_handles) == 1:
                        break
                    else:
                        continue
                else:
                    driver.get(product_url)
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)
            driver.execute_script("window.scrollTo(0, 0);")
            # Scroll to end
            mouse.send_keys(Keys.PAGE_DOWN).perform()
            time.sleep(2)
            driver.execute_script("window.scrollTo(0, 0);")
            try:
                driver.execute_script("document.querySelector('.pdp-overlay_theme_sfo').remove()")
                driver.execute_script("document.querySelector('.delivery__options').remove()")
                driver.execute_script("document.elementFromPoint(50, 50).click()")
                print('Overseas overlay removed')
            except JavascriptException:
                pass
            # Oversea overlay
            try:
                driver.execute_script("document.querySelector('#New_LzdSiteNav').remove()")
                print('Navbar removed')
            except JavascriptException:
                pass
            # Get product name
            product_name = wait.until(
                ec.presence_of_element_located((By.CLASS_NAME, 'pdp-mod-product-badge-title'))).text
            # Add product name to sheet
            worksheet = load_workbook(os.path.join(directory, 'product_info.xlsx'))
            worksheet['Product Name'].append([product_name, ])
            while True:
                try:
                    worksheet.save(os.path.join(directory, 'product_info.xlsx'))
                    break
                except PermissionError:
                    print(RED, 'Close the file', RESET, ' product_info.xlsx')
                    time.sleep(1)
            # Main images
            image_element = wait.until(ec.presence_of_all_elements_located((
                By.CSS_SELECTOR, '.next-slick-track .next-slick-slide .item-gallery__image-wrapper img')))
            if 'https://img.lazcdn.com/g/tps/tfs' in image_element[0].get_attribute('src'):
                image_element.pop(0).click()  # click on YouTube thumbnail
                # Get video url
                try:
                    video_url = wait.until(ec.presence_of_element_located((
                        By.CSS_SELECTOR, '.item-gallery__video-player iframe'))).get_attribute('src')
                except (NoSuchElementException, TimeoutException, StaleElementReferenceException):
                    video_url = wait.until(ec.presence_of_element_located((
                        By.CSS_SELECTOR, '.item-gallery__video-player video'))).get_attribute('src')
                # Load current worksheet and add video URL
                worksheet = load_workbook(os.path.join(directory, 'product_info.xlsx'))
                worksheet['Video URL'].cell(worksheet['Video URL'].max_row + 1, 1).value = video_url.split('?')[0]
                while True:
                    try:
                        worksheet.save(os.path.join(directory, 'product_info.xlsx'))
                        break
                    except PermissionError:
                        print(RED + 'Close the file' + RESET + ' product_info.xlsx')
                        time.sleep(1)
                # Close video dialogue
                wait.until(
                    ec.presence_of_element_located(
                        (By.CSS_SELECTOR, '.item-gallery__video-dialog > a .next-icon'))).click()
            # Get image urls
            for image in image_element:
                image_url = image.get_attribute('src')
                print(image_url)
                # Download image
                download_image(image_url)

            # SKU variation
            try:
                sku_elements = wait.until(ec.presence_of_all_elements_located((
                    By.CSS_SELECTOR, '.sku-selector .sku-prop:nth-child(1) .sku-prop-content > span')))
                for sku_element in sku_elements:
                    if not (sku_element.get_attribute('class') == 'sku-variable-img-wrap-selected' or
                            sku_element.get_attribute('class') == 'sku-variable-name-selected'):
                        sku_element.click()
                    try:
                        image_url = sku_element.find_element(
                            By.CSS_SELECTOR, '.sku-variable-img img').get_attribute('src')
                        download_image(image_url)
                    except NoSuchElementException:
                        print('No image found in SKU')
            except TimeoutException:
                print('SKU section not found')

            # Highlight
            try:
                highlight = wait.until(
                    ec.presence_of_element_located((By.CLASS_NAME, 'pdp-product-highlights'))).text
                # Load current worksheet and add highlight
                worksheet = load_workbook(os.path.join(directory, 'product_info.xlsx'))
                worksheet['Highlights'].cell(worksheet['Highlights'].max_row + 1, 2).value = highlight
                save_sheet(worksheet)
            except TimeoutException:
                print('Highlight not found')

            # Description
            try:
                highlight = wait.until(
                    ec.presence_of_element_located((By.CLASS_NAME, 'detail-content'))).text
                # Load current worksheet and add highlight
                worksheet = load_workbook(os.path.join(directory, 'product_info.xlsx'))
                worksheet['Description'].cell(worksheet['Description'].max_row + 1, 1).value = highlight
                save_sheet(worksheet)
                worksheet['Highlights'].alignment = Alignment(wrap_text=True)
            except TimeoutException:
                print('Highlight not found')
            # Close the tab
            driver.close()
        except Exception as e:
            print(e)
    # Safely close chrome driver
    try:
        driver.close()
    except Exception as e:
        print(e)


def create_xml():
    worksheet = load_workbook(os.path.join(directory, 'product_info.xlsx'))

    # Create the xml file
    # Create the root element
    root = Et.Element("Request")
    product = Et.SubElement(root, "Product")
    # Product attributes
    Et.SubElement(product, "PrimaryCategory").text = str(worksheet['Category']['A1'].value)
    Et.SubElement(product, "SPUId").text = ""
    Et.SubElement(product, "AssociatedSku").text = ""

    # Main images
    main_images = Et.SubElement(product, "Images")
    Et.SubElement(main_images, "Image").text = ''

    # Attributes
    attributes = Et.SubElement(product, "Attributes")
    product_name_eng = worksheet['Product Name']['A1'].value
    Et.SubElement(attributes, "name").text = product_name_eng.title()
    Et.SubElement(attributes, "name_en").text = product_name_eng

    # Self attributes
    worksheet = load_workbook(os.path.join(directory, 'product_info.xlsx'))
    for row in worksheet['Attributes'].iter_rows(min_row=2, values_only=True):
        name, value = row
        Et.SubElement(attributes, name).text = value

    # Highlights
    highlight = '<ul>'
    for h_line in worksheet['Highlights'].iter_rows(min_row=0, values_only=True):
        if h_line[0] is not None:
            highlight += f'<li>{re.sub(r'^[^\w\s]', '', h_line[0])}</li>'
    highlight += '</ul>'
    Et.SubElement(attributes, "short_description").text = highlight
    Et.SubElement(attributes, "short_description_en").text = highlight

    # Description
    Et.SubElement(attributes, "description").text = ''
    Et.SubElement(attributes, "description_en").text = ''
    Et.SubElement(attributes, "promotion_whitebkg_image").text = ''
    Et.SubElement(attributes, "brand").text = 'No Brand'
    Et.SubElement(attributes, "warranty").text = 'N/A'
    Et.SubElement(attributes, "warranty_type").text = 'No Warranty'
    Et.SubElement(attributes, "Hazmat").text = 'None'
    Et.SubElement(attributes, "delivery_option_sof").text = 'No'
    video = worksheet['Video URL']['A1'].value
    if video is not None:
        youtube_video_id = extract_youtube_id(video)
        print(f'Video ID: {youtube_video_id}')
        Et.SubElement(attributes, "video").text = 'https://www.youtube.com/watch?v=' + youtube_video_id
    Et.SubElement(attributes, "disableAttributeAutoFill").text = 'false'

    # SKU attributes
    highest_price = 0
    sku_image_list = []
    skus = Et.SubElement(product, "Skus")
    for row in worksheet['SKU Info'].iter_rows(min_row=2, values_only=True):
        color_family, seller_sku, price, image_url, dimensions, weight = row
        # For multi price
        if color_family.lower() in 'multi':
            price = highest_price * 2
        else:
            highest_price = max(highest_price, int(price))
        image_url = str(image_url)
        if '.' not in image_url:
            image_url = full_source(image_url)
        sku_image_list.append(image_url)
        sku = Et.SubElement(skus, "Sku")
        # Et.SubElement(sku, "SkuId").text = "2120152622"
        Et.SubElement(sku, "SellerSku").text = f'{seller_sku.title()}-{color_family.title()}'
        sale_prop = Et.SubElement(sku, "saleProp")
        Et.SubElement(sale_prop, "color_family").text = color_family.title()
        # Et.SubElement(sale_prop, "size").text = "10"
        Et.SubElement(sku, "quantity").text = "1000"
        Et.SubElement(sku, "price").text = str(max(math.ceil(int(price) * 1.2 / 50) * 50, 200))
        Et.SubElement(sku, "special_from_date").text = "2024-09-18 00:00:00"
        Et.SubElement(sku, "special_to_date").text = "2099-08-05 00:00:00"
        Et.SubElement(sku, "special_price").text = str(price)
        Et.SubElement(sku, "package_length").text = dimensions.split('x')[0]
        Et.SubElement(sku, "package_width").text = dimensions.split('x')[1]
        Et.SubElement(sku, "package_height").text = dimensions.split('x')[2]
        Et.SubElement(sku, "package_weight").text = str(weight)
        Et.SubElement(sku, "package_content").text = f"1 x {product_name_eng}"
        sku_images = Et.SubElement(sku, "Images")
        Et.SubElement(sku_images, "Image").text = ''

    cursor.execute('SELECT access_token FROM open_daraz WHERE shop_name = ?', ('Digital Box',))
    access_token = cursor.fetchone()[0]
    daraz_api = DarazAPI(access_token, 'Digital Box')

    # Set self attributes
    # worksheet = load_workbook(os.path.join(directory, 'product_info.xlsx'))
    for attribute in daraz_api.category_attributes(str(worksheet['Category']['A1'].value))['data']:
        is_key_prop = attribute['advanced']['is_key_prop']
        is_sale_prop = attribute['is_sale_prop']
        name = attribute['name']
        input_type = attribute['input_type']
        attribute_option = []
        if 'options' in attribute:
            for option in attribute['options']:
                attribute_option.append(option['name'])
        is_mandatory = attribute['is_mandatory']
        attribute_type = attribute['attribute_type']
        label = attribute['label']
        # if attribute_type == 'normal' and is_mandatory != 0:
        if attribute_type == 'normal':
            if root.find('Product').find('Attributes').find(name) is None:
                if attribute_option:  # Attribute option
                    print(YELLOW, label + RED + ('*' if is_mandatory != 0 else ''), GREEN, input_type, RESET)
                    for index, option in enumerate(attribute_option):
                        print(str(index + 1) + '. ', option)
                    user_input = input(f'→ {CYAN} {'Key Prop' if is_key_prop == 1 else ''} {RESET}')
                    multi_attribute = ''
                    if user_input != '':
                        for input_key in user_input.split(' '):
                            multi_attribute += attribute_option[int(input_key) - 1] + ','
                        Et.SubElement(attributes, name).text = multi_attribute.strip(',')
                        worksheet['Attributes'].append([name, multi_attribute.strip(',')])
                    else:
                        if is_mandatory != 0:
                            print(RED + 'Mandatory field' + RESET)
                            raise Exception('Mandatory field')
                else:  # Without attribute option
                    print(YELLOW, label + RED + ('*' if is_mandatory != 0 else ''), GREEN, input_type, RESET)
                    user_input = input(f'→ {CYAN} {'Key Prop' if is_key_prop == 1 else ''} {RESET}')
                    if user_input != '':
                        Et.SubElement(attributes, name).text = user_input
                        worksheet['Attributes'].append([name, user_input.capitalize()])
                # Check if the attribute is a key prop
                if is_sale_prop != 0:
                    print(RED + 'Sale Prop:', name + RESET)
    save_sheet(worksheet)
    return root, sku_image_list


def main():
    winsound.Beep(2000, 800)
    if input('Scrap product (y/n)? ').lower() == 'y':
        setup_directory()
        setup_excel()
        scrap_handler()
    xml_root, sku_image_list = create_xml()

    # Upload to daraz
    cursor.execute('SELECT shop_name, access_token FROM open_daraz LIMIT -1 OFFSET 0')

    argument = [(xml_root, sku_image_list, shop_name) for shop_name, access_token in cursor.fetchall()]
    import multiprocessing
    with multiprocessing.Pool(processes=6) as pool:
        pool.starmap(upload_product, argument)

    winsound.Beep(1000, 100)
    print('Upload completed')


if __name__ == '__main__':
    main()
