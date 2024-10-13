import re
import time
import gspread
import sqlite3
import requests
import random
from PIL import Image
from io import BytesIO
import pandas as pd
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment
from oauth2client.service_account import ServiceAccountCredentials
from gimini import model

# connect to database
conn = sqlite3.connect('order_details.db')
cursor = conn.cursor()
cursor_view = sqlite3.connect('order_details.db').cursor()
cursor.execute('DELETE FROM new_product WHERE TRUE')
conn.commit()

# Help image link
for serial, shop_name in cursor.execute('SELECT serial, shop_name FROM unique_image').fetchall():
    print(f'{serial}. {shop_name}')
# cursor.execute('SELECT url FROM unique_image WHERE serial = ?', (int(input('Enter serial: ')),))
cursor.execute('SELECT url FROM unique_image WHERE serial = ?', (12,))
help_image = cursor.fetchone()[0]


def column(header_name):
    return df.columns.get_loc(header_name) + 1


def cdn_link(link: str):
    if link == '':
        return link
    full_link = link.split('_')[0].strip()
    if 'img.drz.lazcdn.com/static/bd/p/' in full_link:
        image_code = full_link.split('static/bd/p/')[1]
        link = 'https://static-01.daraz.com.bd/p/' + image_code
    elif 'static-01.daraz.com.bd/p/' in full_link:
        link = full_link
    else:
        print(f'New link: {link}')
        link = full_link
    return link + f'?q={str(time.time()).replace(".", "")}'


def image_dimensions(url):
    # request the image from the URL until it is successfully fetched
    while True:
        try:
            response = requests.get(url)
            break
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout):
            time.sleep(5)
    img = Image.open(BytesIO(response.content))
    return img.size  # returns (width, height)


def add_image_element():
    (image_width, image_height) = image_dimensions(image_url)
    if image_width < 928:
        image_width = 928
        image_height = image_height / image_width * 928
    image_element = soup.new_tag('img')
    image_element['style'] = f"width:{image_width}px;height:{image_height}px;display:inline;vertical-align:middle"
    image_element['src'] = image_url
    paragraph.append(image_element)


def insert_link(row, column, link):
    cell = excel_sheet.cell(row=row, column=column)
    cell.value = link
    # cell.hyperlink = link
    # cell.style = "Hyperlink"


# def copy_cell(source_cell, target_cell):
#     target_cell.value = source_cell.value  # Copy value
#     if source_cell.has_style:  # Check if the source cell has formatting
#         target_cell.font = source_cell.font  # Copy font
#         target_cell.border = source_cell.border  # Copy border
#         target_cell.fill = source_cell.fill  # Copy fill (background color)
#         target_cell.alignment = source_cell.alignment  # Copy alignment
#         target_cell.number_format = source_cell.number_format  # Copy number format

# Define the scope
scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]

# Add the path to the downloaded JSON key file
creds = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scope)

# Authorize the client
client = gspread.authorize(creds)

# Open the Google Sheet
spreadsheet = client.open('Daraz Product')

# Variable
file_path = r'D:\Downloads\Phone accessory.xlsx'
excel = load_workbook(file_path, rich_text=True)
df = pd.read_excel(file_path, sheet_name='PhoneCases', header=1)
excel_sheet = excel['PhoneCases']
sheet = spreadsheet.worksheet('Phone Case')

print('Creating new product records...', end='')
sel_row = 3
# Update records
product = None
for index, row in enumerate(sheet.get_all_records()):
    # Product name
    product_name_eng = row['Product name (English)'].title()
    cursor_view.execute('SELECT * FROM new_product WHERE title = ?', (product_name_eng,))
    if cursor_view.fetchall():
        continue
    sku_image = row['SKU Image']
    if product != product_name_eng:
        print(f'\n{index + 1} ', end='')
        conn.commit()
        excel.save(file_path)
        while True:
            try:
                response = model.generate_content(
                    [f":Write me one new SEO friendly title based on: {product_name_eng}",
                     'If necessary you can add new words to the title to make it more SEO friendly',
                     'Only write title, don\'t add extra information'])
                auto_text = response.text.replace('#', '').replace('*', '').strip()
                break
            except Exception as e:
                print(e)
                if '429 Resource has been exhausted (e.g. check quota).' in str(e):
                    time.sleep(30)
                time.sleep(5)
        if '\n' in auto_text or auto_text == '':
            auto_title = input(f"Enter product name for {product_name_eng}({response.text}): ")
        else:
            auto_title = auto_text
        main_image_series = list(range(1, 8))
        random.shuffle(main_image_series)
        product = product_name_eng
        unique_ext = ' QR:' + time.strftime('%y%m%d%H%M%S')
        seller_sku = row['Seller SKU'].title()
        if seller_sku == '':
            while True:
                try:
                    response = model.generate_content(
                        ["Give me only one product identification text of product title:"
                         f" {product_name_eng}"])
                    auto_text = response.text.replace('#', '').replace('*', '').strip()
                    break
                except Exception as e:
                    print(e)
                    if '429 Resource has been exhausted (e.g. check quota).' in str(e):
                        time.sleep(30)
                    time.sleep(5)
            if '\n' in auto_text or auto_text == '':
                seller_sku = input(f"Enter Seller SKU for {product_name_eng}: ")
            else:
                seller_sku = auto_text
        cursor.execute('SELECT * FROM new_product WHERE seller_sku = ?', (seller_sku,))
        if cursor.fetchall():
            seller_sku = seller_sku + '~' + random.randint(100, 999).__str__()
        cursor.execute('INSERT INTO new_product VALUES (?, ?, ?)', (product_name_eng, seller_sku, index + 1))

        # Main image, Description
        soup = BeautifulSoup('', 'html.parser')
        html = soup.new_tag('article')
        html['style'] = "white-space:break-spaces"
        html['class'] = "lzd-article"
        soup.append(html)
        paragraph = soup.new_tag('p')
        paragraph['style'] = ("line-height:1.7;text-align:center;text-indent:0;margin-left:0;margin-top:0;margin"
                              "-bottom:0")
        html.append(paragraph)
        text_element = soup.new_tag('span')
        text_element['style'] = ("font-weight:bold;color:rgb(64, 64, 64);"
                                 "background-color:rgb(255, 217, 102);font-size:12pt")
        text_element.string = row['Main Description']
        paragraph.append(text_element)

        for serial, image_series in enumerate(main_image_series):
            image_url = row[f'Product Images{image_series + 1}'].strip()
            if image_url == 'https://static-01.daraz.com.bd/p/132f1553877bbf8f0314d2cddd3a401c.png':
                image_url = help_image
            if image_url == '':
                continue
            image_url = cdn_link(image_url)
            insert_link(sel_row, column(f'Product Images{serial + 2}'), image_url)
            add_image_element()
            print('* ', end='')
    else:
        # Copy previous row to current row
        for col in range(4, 13):
            excel_sheet.cell(row=sel_row, column=col, value=excel_sheet.cell(row=sel_row - 1, column=col).value)
            # copy_cell(excel_sheet.cell(row=sel_row - 1, column=col), excel_sheet.cell(row=sel_row, column=col))
    excel_sheet.cell(row=sel_row, column=column('Group No'), value=product_name_eng)
    excel_sheet.cell(row=sel_row, column=column('*Product Name(English)'), value=auto_title + unique_ext)
    product_name_bn = row['Product name (Bangla)'] + unique_ext
    excel_sheet.cell(row=sel_row, column=column('Product Name(Bengali) look function'), value=product_name_bn)
    print('* ', end='')
    # Image, Description, SKU
    image_url = cdn_link(row[f'Product Images1'])
    insert_link(sel_row, column('*Product Images1'), image_url)
    insert_link(sel_row, column('White Background Image'), image_url)
    # SKU, Price
    color_family = row['Color Family'].title()
    try:
        excel_sheet.cell(row=sel_row, column=column('*Color Family'), value=color_family)
    except KeyError:
        excel_sheet.cell(row=sel_row, column=column('Color Family'), value=color_family)
    excel_sheet.cell(row=sel_row, column=column('SellerSKU'), value=seller_sku + '-' + color_family)
    sku_image = row['SKU Image']
    insert_link(sel_row, column('Images1'), cdn_link(sku_image))
    excel_sheet.cell(row=sel_row, column=column('*Quantity'), value='1000')
    price = int(row['Price']) + 20
    excel_sheet.cell(row=sel_row, column=column('*Price'), value=str(max(round((price * 1.5) / 50) * 50, 120)))
    excel_sheet.cell(row=sel_row, column=column('SpecialPrice'), value=str(price))
    excel_sheet.cell(row=sel_row, column=column('SpecialPrice Start'),
                     value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    excel_sheet.cell(row=sel_row, column=column('SpecialPrice End'), value='2099-08-05 00:00:00')

    # Update the main description
    excel_sheet.cell(row=sel_row, column=column('Main Description'), value=str(soup))
    # Additional
    excel_sheet.cell(row=sel_row, column=column('*Brand'), value='No Brand')
    excel_sheet.cell(row=sel_row, column=column('Warranty'), value='N/A')
    excel_sheet.cell(row=sel_row, column=column('Warranty Type'), value='No Warranty')
    # Measurement, weight
    weight = row['Weight']
    excel_sheet.cell(row=sel_row, column=column('*Package Weight (kg)'), value=weight)
    length, width, height = row['Dimension (cm)'].split('x')
    excel_sheet.cell(row=sel_row, column=column('*Package Length(cm) * Width(cm) * Height(cm)-Length (cm)'),
                     value=length)
    excel_sheet.cell(row=sel_row, column=column('*Package Length(cm) * Width(cm) * Height(cm)-Width (cm)'),
                     value=width)
    excel_sheet.cell(row=sel_row, column=column('*Package Length(cm) * Width(cm) * Height(cm)-Height (cm)'),
                     value=height)
    excel_sheet.cell(row=sel_row, column=column('Dangerous Goods'), value='None')
    excel_sheet.cell(row=sel_row, column=column("*What's in the box"), value='1 X ' + product_name_eng)
    # Attribute
    # attributes = row['Attributes']
    # if attributes != '':
    #     attributes = eval(attributes)
    #     for attribute in attributes:
    #         excel_sheet.cell(row=sel_row, column=column(attribute), value=attributes[attribute])
    # Highlights
    soup = BeautifulSoup('', 'html.parser')
    html = soup.new_tag('article')
    html['style'] = "white-space:break-spaces"
    html['class'] = "lzd-article"
    soup.append(html)
    unordered = soup.new_tag('ul')
    html.append(unordered)
    for highlight in row['Highlights'].split('\n'):
        unordered_list = soup.new_tag('li')
        unordered.append(unordered_list)
        line = soup.new_tag('div')
        line['style'] = "line-height:1.7;text-align:left;text-indent:0"
        unordered_list.append(line)
        line_str = soup.new_tag('span')
        line_str.string = highlight.replace('●', '').replace('•', ' ').strip()
        line.append(line_str)
    excel_sheet.cell(row=sel_row, column=df.columns.get_loc('Highlights') + 1, value=str(soup))
    # format row as wrap text
    for cell in excel_sheet[f'{sel_row}:{sel_row}']:
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
    excel_sheet.row_dimensions[sel_row].height = 100
    sel_row += 1
    # if sel_row > 6:
    #     break
